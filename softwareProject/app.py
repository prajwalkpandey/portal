from flask import Flask,jsonify,request,render_template,redirect,url_for,session
import os
import openpyxl,json
from flask_login import LoginManager,login_required, login_user, logout_user
from flask_sqlalchemy import SQLAlchemy
from flask_bcrypt import Bcrypt
#from flask_login import LoginManager
#from login import User
#import login

students=openpyxl.load_workbook(filename="Students.xlsx")
transactions=openpyxl.load_workbook(filename="Transaction.xlsx")
courses=openpyxl.load_workbook(filename="Courses.xlsx")
backlogs=openpyxl.load_workbook(filename="Backlogs.xlsx")
registrations=openpyxl.load_workbook(filename="Registrations.xlsx")
raised=openpyxl.load_workbook(filename="raisedCases.xlsx")

app = Flask(__name__)
login_manager=LoginManager()
db = SQLAlchemy()
bcrypt=Bcrypt(app)

@app.route('/')
def index():
    return jsonify({"data":"Hihahu"})

@app.route('/student')
def roll():
    data=request.json
    rollno=(data['rollno'])
    return jsonify(student(rollno))

@app.route('/courses')
def course():
    data=request.json
    rollno=(data['rollno'])
    semester=(data['semester'])
    resp=fetchCourses(semester)
    blogs=checkBacklogs(rollno)
    return ({"courses":resp,"backlogs":blogs})

@app.route('/registration',methods=['GET', 'POST'])
def completeRegistration():
    if request.method == 'POST':
        data=request.json
        return registrationSave(data)
    if request.method == 'GET':
        data = request.json
        if(data == None):
            return jsonify(fetchRegistration())
        else:
            return jsonify(fetchGrantToken(data['rollno']))
        
def fetchGrantToken(rollno):
    registration_obj = registrations.active
    max_row =  registration_obj.max_row
    for i in range(2,max_row+1):
        registration_cell = registration_obj.cell(row=i,column=1)
        if(registration_cell.value == rollno):
            data ={
                "grantToken":registration_obj.cell(row=i,column=6).value
            }
            return json.dumps(data)

def fetchRegistration():
    registration_obj = registrations.active
    max_row =  registration_obj.max_row
    data = []
    for i in range(2,max_row+1):
        dat = {
            "rollno":registration_obj.cell(row=i,column=1).value,
            "name":registration_obj.cell(row=i,column=2).value,
            "branch":registration_obj.cell(row=i,column=3).value,
            "semester":registration_obj.cell(row=i,column=4).value,
            "electives":registration_obj.cell(row=i,column=5).value,
            "token":registration_obj.cell(row=i,column=6).value,
            "status":registration_obj.cell(row=i,column=7).value,
            "fine":registration_obj.cell(row=i,column=8).value,
        }
        data.append(dat);
    return json.dumps(data);

def registrationSave(data):
    rollno=(data['rollno'])
    semester=(data['semester'])
    name=data['name']
    branch=data['branch']
    electives=data['electives']
    payment=data['status']
    fine=data['fine']
    
    registration_obj = registrations.active
    max_row =  registration_obj.max_row
    registration_obj.cell(row=max_row+1,column=1).value=rollno
    registration_obj.cell(row=max_row+1,column=2).value=name
    registration_obj.cell(row=max_row+1,column=3).value=branch
    registration_obj.cell(row=max_row+1,column=4).value=semester
    registration_obj.cell(row=max_row+1,column=5).value=electives
    registration_obj.cell(row=max_row+1,column=6).value=str(rollno)+branch
    registration_obj.cell(row=max_row+1,column=7).value=payment
    registration_obj.cell(row=max_row+1,column=8).value=fine
    registrations.save(filename="/home/prakhar/Documents/Registrations.xlsx")
    return jsonify({"saved":True})
    
def registeredQuery():
    # rollno=(data['rollno'])
    # semester=(data['semester'])
    # name=data['name']
    # branch=data['branch']
    # electives=data['electives']
    # payment=data['status']
    # fine=data['fine']
    obj={}
    registration_obj = registrations.active
    max_row =  registration_obj.max_row
    for i in range(1,max_row):
        temp={}
        temp['rollno']=registration_obj.cell(row=i,column=1).value
        temp['name']=registration_obj.cell(row=i,column=2).value
        temp['branch']=registration_obj.cell(row=i,column=3).value
        temp['semester']=registration_obj.cell(row=i,column=4).value
        temp['electives']=registration_obj.cell(row=i,column=5).value
        # temp['electives']=registration_obj.cell(row=i,column=6).value
        temp['payment']=registration_obj.cell(row=i,column=7).value
        temp['fine']=registration_obj.cell(row=i,column=8).value
        obj[i]=temp
    # registrations.save(filename="/home/prakhar/Documents/Registrations.xlsx")
    return obj #jsonify({"saved":True})
    
def checkBacklogs(rollno):
    backlog_obj = backlogs.active
    max_col =  backlog_obj.max_column
    max_row =  backlog_obj.max_row
    print(max_row)
    data = []
    for i in range(2,max_row+1):
        course_cell = backlog_obj.cell(row=i,column=1)
        if(course_cell.value == rollno):
            dat = {
                "id":backlog_obj.cell(row=i,column=2).value,
                "name":backlog_obj.cell(row=i,column=3).value,         
            }
            data.append(dat)
    return json.dumps(data)
    
    
def fetchCourses(semester):
    course_obj=courses.active
    max_col =  course_obj.max_column
    max_row =  course_obj.max_row
    data = []
    for i in range(2,max_row+1):
        course_cell = course_obj.cell(row=i,column=1)
        if(course_cell.value == semester):
            dat = {
                "id":course_obj.cell(row=i,column=2).value,
                "name":course_obj.cell(row=i,column=3).value,
                "type":course_obj.cell(row=i,column=4).value,
                "faculty":course_obj.cell(row=i,column=5).value,           
            }
            data.append(dat)
    return json.dumps(data)

def student(rollno):
    # print("Student",rollno)
    student_obj= students.active
    max_col =  student_obj.max_column
    max_row =  student_obj.max_row
    for i in range(2,max_row+1):
        student_cell = student_obj.cell(row=i,column=1)
        if(student_cell.value == rollno):
            std={
                "name":student_obj.cell(row=i,column=2).value,
                "email":student_obj.cell(row=i,column=3).value,
                "branch":student_obj.cell(row=i,column=4).value,
                "year":student_obj.cell(row=i,column=5).value,
                "category":student_obj.cell(row=i,column=6).value,
                "laptop":student_obj.cell(row=i,column=7).value,
            }
            std1=json.dumps(std)
            return transaction(std1,rollno)

def fetchCases():
    # print("Student",rollno)
    obj= raised.active
    max_col =  obj.max_column
    max_row =  obj.max_row
    std={}
    for i in range(1,max_row+1):
        std[i]=obj.cell(row=i,column=1).value
    print(std)
    return std

def fetchCase(rollno):
    obj= raised.active
    max_col =  obj.max_column
    max_row =  obj.max_row
    std={}
    for i in range(2,max_row+1):
        caseN = obj.cell(row=i,column=1)
        if(caseN.value == rollno):
            std[rollno]=obj.cell(row=i,column=2).value
            print(std)
            return std

def addCase(data):
    rollno=data['rollno']
    issue=data['case']
    obj = raised.active
    max_row =  obj.max_row
    obj.cell(row=max_row+1,column=1).value=rollno
    obj.cell(row=max_row+1,column=2).value=issue
    raised.save(filename="raisedCases.xlsx")
    print("added")
    return jsonify({"saved":True})            


def transaction(data,rollno):
    transaction_obj = transactions.active
    max_row =  transaction_obj.max_row
    
    for i in range(2,max_row+1):
        transaction_cell=transaction_obj.cell(row=i,column=1)
        if(transaction_cell.value == rollno):
            dat ={
                "student":data,
                "transactionId":transaction_obj.cell(row=i,column=2).value
            }
            return json.dumps(dat)


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == 'GET':
        return render_template("login.html")#, name='name')
    else:
    #    form = LoginForm()
    #    if form.validate_on_submit():
    #    print(request)
     #   print(request.form['username'])
#        print(request.form['password'])
    ##        user = User.query.get(form.email.data)
    ##        if user:
#        if request.form['email'] 
        if bcrypt.check_password_hash('$2b$12$T8bxlwiQAGVXbduNXpGfxe2CiOC7Y5y.gicLykSq4rnz6AMLDPjne', request.form['password']):
#            print(request.form['password'])
#    #                user.authenticated = True
#    #                db.session.add(user)
#    #                db.session.commit()
            session['logged_in'] = True
#                    login_user(user, remember=False)
            return redirect(url_for("admin"))
        else:
            print("FAIL")
        return redirect("/")

@app.route("/admin", methods=["GET","POST"])
#@login_required
def admin():
    if request.method== 'GET':
        if session.get('logged_in'):
            # registrationSaved=registeredQuery()
            # print(registrationSaved)
            # print("=======================")
            allCases=fetchCases()
            print("=======================")
            fetchCase(987654)
            return render_template("admin.html",data=allCases)
           #  return "<html>redirect</html>"
        else:
            return "<html> login</html>"
            print("login")
    # else:
    #     request.form['password']

@app.route("/cases", methods=["GET","POST"])
#@login_required
def cases():
    if request.method== 'GET':
        return render_template("raise.html")
    else:
        newCase=request.form
        print(newCase)
        addCase(newCase)
        # allCases=fetchCases()
        return jsonify({"saved":True})

if __name__ == '__main__':
    app.secret_key = os.urandom(24)
    app.run(debug=True)

class User(db.Model):
    __tablename__ = 'user'

    email = db.Column(db.String, primary_key=True)
    password = db.Column(db.String)
    authenticated = db.Column(db.Boolean, default=False)

    def is_active(self):
        return True

    def get_id(self):
        return self.email

    def is_authenticated(self):
        return self.authenticated

    def is_anonymous(self):
        return False

@login_manager.user_loader
def user_loader(user_id):
    return User.query.get(user_id)
